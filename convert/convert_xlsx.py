#!/usr/bin/env python
"""Convert XLSX files from the BFS statistical atlas into TSV files.
Usage:
  convert_xlsx.py <xlsx_file>
  convert_xlsx.py (-h | --help)
  convert_xlsx.py --version
Options:
  -h --help         Show this screen.
  --version         Show version.
"""
import csv
import sys
from docopt import docopt
from openpyxl import load_workbook


HEADER_OFFSET = 3
FOOTER_OFFSET = 11


def find_data(xlsx_file):
    wb = load_workbook(xlsx_file)
    ws = wb.active
    for row_idx in range(HEADER_OFFSET, ws.max_row - FOOTER_OFFSET):
        row = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row.append(cell.value)
        yield row


if __name__ == '__main__':
    args = docopt(__doc__, version='0.1')
    xlsx_file = args.get('<xlsx_file>')
    rows = list(find_data(xlsx_file))

    header_row = rows[0]
    header_row[0] = 'Regions-ID'
    header_row[1] = 'Regionsname'

    data_rows = [r for r in rows[5:] if r[0] is not None]

    writer = csv.writer(sys.stdout, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(header_row)
    for row in data_rows:
        writer.writerow(row)
